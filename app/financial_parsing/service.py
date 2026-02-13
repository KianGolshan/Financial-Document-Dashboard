import base64
import json
import logging
import tempfile
from pathlib import Path

import fitz  # PyMuPDF
import anthropic
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from sqlalchemy.orm import Session

from app.config import settings
from app.database import SessionLocal
from app.documents.models import Document
from app.exceptions import bad_request, not_found
from app.financial_parsing.models import EditLog, FinancialStatement, LineItem, ParseJob
from app.investments.models import Investment

logger = logging.getLogger(__name__)

EXTRACTION_PROMPT = """\
You are a financial statement extraction engine. Extract structured data from the following document pages.

For each financial statement found (income statement, balance sheet, cash flow statement), output a JSON object.

Return a JSON array of statement objects. Each object must have:
- "statement_type": one of "income_statement", "balance_sheet", "cash_flow"
- "period": the reporting period as written (e.g., "Year Ended December 31, 2023", "Q3 2023")
- "period_end_date": ISO date if determinable (e.g., "2023-12-31"), or null
- "currency": currency code (e.g., "USD"), or null
- "unit": unit of measure as stated (e.g., "in thousands", "in millions"), or null
- "source_pages": page numbers where this statement was found (e.g., "5-7")
- "line_items": array of line item objects

Each line_item must have:
- "category": a canonical category from the lists below
- "label": the exact label as printed in the document
- "value": numeric value (parentheses mean negative, e.g., "(500)" = -500). null if no value.
- "is_total": true if this is a total/subtotal line
- "indent_level": 0 for top-level, 1 for indented, 2 for double-indented, etc.

Canonical categories for income_statement:
revenue, cost_of_revenue, gross_profit, operating_expenses, research_and_development,
selling_general_admin, depreciation_amortization, operating_income, interest_expense,
interest_income, other_income_expense, income_before_tax, income_tax, net_income,
earnings_per_share, other

Canonical categories for balance_sheet:
cash_and_equivalents, short_term_investments, accounts_receivable, inventory,
other_current_assets, total_current_assets, property_plant_equipment, goodwill,
intangible_assets, long_term_investments, other_non_current_assets, total_assets,
accounts_payable, short_term_debt, accrued_liabilities, other_current_liabilities,
total_current_liabilities, long_term_debt, other_non_current_liabilities,
total_liabilities, common_stock, retained_earnings, treasury_stock,
other_equity, total_stockholders_equity, total_liabilities_and_equity, other

Canonical categories for cash_flow:
net_income, depreciation_amortization, stock_based_compensation,
changes_in_working_capital, other_operating, operating_cash_flow,
capital_expenditures, acquisitions, purchases_of_investments,
sales_of_investments, other_investing, investing_cash_flow,
debt_issued, debt_repaid, shares_issued, shares_repurchased,
dividends_paid, other_financing, financing_cash_flow,
net_change_in_cash, beginning_cash, ending_cash, other

If a document contains multiple periods (e.g., current year and prior year side by side),
create a separate statement object for each period.

Preserve the document's ordering of line items in the sort_order (use sequential integers starting from 0).

IMPORTANT: Return ONLY the JSON array. No markdown, no explanation, no code fences.
If no financial statements are found in these pages, return an empty array: []
"""


def _extract_chunks(pdf_path: str) -> list[dict]:
    """Split PDF into overlapping page-window chunks.
    Each chunk contains page images (as base64 PNG) and any extractable text."""
    chunk_size = settings.PARSING_CHUNK_SIZE
    overlap = settings.PARSING_CHUNK_OVERLAP

    doc = fitz.open(pdf_path)
    total_pages = len(doc)
    chunks = []

    start = 0
    while start < total_pages:
        end = min(start + chunk_size, total_pages)
        images = []
        text_parts = []
        has_text = False
        for page_num in range(start, end):
            page = doc[page_num]
            # Render page to PNG image
            pix = page.get_pixmap(dpi=150)
            img_bytes = pix.tobytes("png")
            images.append({
                "page_num": page_num + 1,
                "b64": base64.standard_b64encode(img_bytes).decode("ascii"),
            })
            # Also try text extraction
            page_text = page.get_text().strip()
            if page_text:
                has_text = True
                text_parts.append(f"--- Page {page_num + 1} ---\n{page_text}")

        chunks.append({
            "start_page": start + 1,
            "end_page": end,
            "images": images,
            "text": "\n\n".join(text_parts) if has_text else "",
            "has_text": has_text,
        })
        if end >= total_pages:
            break
        start = end - overlap

    doc.close()
    return chunks


def _call_claude(chunk: dict, client: anthropic.Anthropic) -> list[dict]:
    """Send chunk to Claude using images (always) + text (when available)."""
    # Build content blocks: images first, then the prompt
    content = []
    for img in chunk["images"]:
        content.append({
            "type": "image",
            "source": {
                "type": "base64",
                "media_type": "image/png",
                "data": img["b64"],
            },
        })

    # Build prompt — include text if available as supplementary
    prompt_text = EXTRACTION_PROMPT
    if chunk["has_text"]:
        prompt_text += "\n\nExtracted text (may be incomplete, use images as primary source):\n"
        prompt_text += chunk["text"]

    content.append({"type": "text", "text": prompt_text})

    response = client.messages.create(
        model=settings.ANTHROPIC_MODEL,
        max_tokens=8192,
        messages=[{"role": "user", "content": content}],
    )
    raw = response.content[0].text.strip()
    # Strip markdown code fences if present
    if raw.startswith("```"):
        lines = raw.split("\n")
        lines = lines[1:]  # drop opening fence
        if lines and lines[-1].strip() == "```":
            lines = lines[:-1]
        raw = "\n".join(lines)
    if not raw:
        return []
    return json.loads(raw)


def _merge_statements(all_results: list[list[dict]]) -> list[dict]:
    """Deduplicate statements by (type, period), keeping the version with more line items."""
    best: dict[tuple, dict] = {}
    for chunk_results in all_results:
        for stmt in chunk_results:
            key = (stmt["statement_type"], stmt["period"])
            existing = best.get(key)
            if existing is None or len(stmt.get("line_items", [])) > len(existing.get("line_items", [])):
                best[key] = stmt
    return list(best.values())


def run_parsing(job_id: int, pdf_path: str, chunks: list[dict]):
    """Background task that processes chunks and writes results to DB.
    Runs synchronously in a background thread via FastAPI BackgroundTasks."""
    db = SessionLocal()
    try:
        job = db.query(ParseJob).get(job_id)
        if not job:
            return

        client = anthropic.Anthropic(api_key=settings.ANTHROPIC_API_KEY)
        all_results: list[list[dict]] = []
        errors: list[str] = []

        for chunk in chunks:
            try:
                result = _call_claude(chunk, client)
                all_results.append(result)
            except Exception as e:
                logger.error("Chunk %d-%d failed: %s", chunk["start_page"], chunk["end_page"], e)
                errors.append(f"Pages {chunk['start_page']}-{chunk['end_page']}: {e}")

            # Update progress
            job = db.query(ParseJob).get(job_id)
            if job:
                job.completed_chunks = job.completed_chunks + 1
                db.commit()

        if not all_results:
            job = db.query(ParseJob).get(job_id)
            job.status = "failed"
            job.error_message = "; ".join(errors) if errors else "No results extracted"
            db.commit()
            return

        merged = _merge_statements(all_results)

        # Delete any old statements for this document (re-parse support)
        old = db.query(FinancialStatement).filter(
            FinancialStatement.document_id == job.document_id
        ).all()
        for s in old:
            db.delete(s)
        db.flush()

        # Write new statements
        for stmt_data in merged:
            stmt = FinancialStatement(
                document_id=job.document_id,
                statement_type=stmt_data["statement_type"],
                period=stmt_data["period"],
                period_end_date=stmt_data.get("period_end_date"),
                currency=stmt_data.get("currency"),
                unit=stmt_data.get("unit"),
                source_pages=stmt_data.get("source_pages"),
                raw_response=json.dumps(stmt_data),
            )
            db.add(stmt)
            db.flush()

            for idx, item_data in enumerate(stmt_data.get("line_items", [])):
                li = LineItem(
                    statement_id=stmt.id,
                    category=item_data.get("category", "other"),
                    label=item_data.get("label", ""),
                    value=item_data.get("value"),
                    is_total=item_data.get("is_total", False),
                    indent_level=item_data.get("indent_level", 0),
                    sort_order=item_data.get("sort_order", idx),
                )
                db.add(li)

        job = db.query(ParseJob).get(job_id)
        job.status = "completed"
        if errors:
            job.error_message = "Partial errors: " + "; ".join(errors)
        db.commit()

    except Exception as e:
        logger.exception("Parsing job %d failed", job_id)
        db.rollback()
        try:
            job = db.query(ParseJob).get(job_id)
            if job:
                job.status = "failed"
                job.error_message = str(e)
                db.commit()
        except Exception:
            pass
    finally:
        db.close()


def parse_document_financials(db: Session, investment_id: int, document_id: int) -> tuple[ParseJob, list[dict]]:
    """Validate and create a parse job. Returns (job, chunks) for the caller to schedule."""
    if not settings.ANTHROPIC_API_KEY:
        raise bad_request("ANTHROPIC_API_KEY is not configured")

    doc = db.query(Document).filter(
        Document.id == document_id,
        Document.investment_id == investment_id,
    ).first()
    if not doc:
        raise not_found("Document not found")

    if doc.document_type.lower() != ".pdf":
        raise bad_request("Only PDF documents can be parsed for financial statements")

    pdf_path = Path(doc.file_path)
    if not pdf_path.exists():
        raise not_found("PDF file not found on disk")

    # Check for active job
    active = db.query(ParseJob).filter(
        ParseJob.document_id == document_id,
        ParseJob.status.in_(["pending", "processing"]),
    ).first()
    if active:
        raise bad_request("A parsing job is already in progress for this document")

    # Extract chunks synchronously (fast with PyMuPDF)
    chunks = _extract_chunks(str(pdf_path))

    job = ParseJob(
        document_id=document_id,
        status="processing",
        total_chunks=len(chunks),
        completed_chunks=0,
    )
    db.add(job)
    db.commit()
    db.refresh(job)

    return job, chunks


def get_parse_job(db: Session, document_id: int) -> ParseJob | None:
    """Get the most recent parse job for a document."""
    return db.query(ParseJob).filter(
        ParseJob.document_id == document_id
    ).order_by(ParseJob.created_at.desc()).first()


def get_statements_for_document(db: Session, document_id: int) -> list[FinancialStatement]:
    """Get all financial statements for a document."""
    return db.query(FinancialStatement).filter(
        FinancialStatement.document_id == document_id
    ).order_by(FinancialStatement.statement_type, FinancialStatement.period).all()


def get_statement(db: Session, statement_id: int) -> FinancialStatement:
    """Get a single financial statement with line items."""
    stmt = db.query(FinancialStatement).filter(
        FinancialStatement.id == statement_id
    ).first()
    if not stmt:
        raise not_found("Financial statement not found")
    return stmt


def delete_financials(db: Session, document_id: int):
    """Delete all parsed financial data for a document."""
    db.query(FinancialStatement).filter(
        FinancialStatement.document_id == document_id
    ).delete()
    db.query(ParseJob).filter(
        ParseJob.document_id == document_id
    ).delete()
    db.commit()


def export_to_excel(db: Session, document_id: int) -> str:
    """Export financial statements to an Excel workbook. Returns temp file path."""
    statements = get_statements_for_document(db, document_id)
    if not statements:
        raise not_found("No financial statements found for this document")

    wb = Workbook()
    # Remove default sheet
    wb.remove(wb.active)

    type_labels = {
        "income_statement": "Income Statement",
        "balance_sheet": "Balance Sheet",
        "cash_flow": "Cash Flow",
    }

    # Group by statement type
    by_type: dict[str, list[FinancialStatement]] = {}
    for stmt in statements:
        by_type.setdefault(stmt.statement_type, []).append(stmt)

    for stmt_type, stmts in by_type.items():
        sheet_name = type_labels.get(stmt_type, stmt_type)[:31]
        ws = wb.create_sheet(title=sheet_name)

        # Header row
        headers = ["Line Item"]
        for stmt in stmts:
            headers.append(stmt.period)
        ws.append(headers)

        # Style header
        for col_idx, _ in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.font = Font(bold=True, size=11)
            cell.alignment = Alignment(horizontal="center")

        ws.cell(row=1, column=1).alignment = Alignment(horizontal="left")

        # Use the statement with the most line items as the template for row ordering
        template_stmt = max(stmts, key=lambda s: len(s.line_items))

        # Build value lookup using display values (edited if available)
        # Key by (stmt_id, display_label) -> display_value
        value_map: dict[tuple[int, str], float | None] = {}
        for stmt in stmts:
            for li in stmt.line_items:
                display_lbl = li.edited_label if li.edited_label is not None else li.label
                display_val = li.edited_value if li.edited_value is not None else li.value
                value_map[(stmt.id, display_lbl)] = display_val

        for li in template_stmt.line_items:
            row_data = []
            indent = "  " * li.indent_level
            display_lbl = li.edited_label if li.edited_label is not None else li.label
            row_data.append(f"{indent}{display_lbl}")
            for stmt in stmts:
                val = value_map.get((stmt.id, display_lbl))
                row_data.append(val)
            ws.append(row_data)

            row_num = ws.max_row
            # Bold totals
            if li.is_total:
                for col_idx in range(1, len(headers) + 1):
                    ws.cell(row=row_num, column=col_idx).font = Font(bold=True)

            # Number formatting for value columns
            for col_idx in range(2, len(headers) + 1):
                cell = ws.cell(row=row_num, column=col_idx)
                cell.number_format = '#,##0.00'

        # Footnote
        footnote_parts = []
        for stmt in stmts:
            if stmt.currency:
                footnote_parts.append(f"Currency: {stmt.currency}")
                break
        for stmt in stmts:
            if stmt.unit:
                footnote_parts.append(f"Unit: {stmt.unit}")
                break
        if footnote_parts:
            ws.append([])
            ws.append([" | ".join(footnote_parts)])

        # Auto-width columns
        for col in ws.columns:
            max_len = 0
            col_letter = col[0].column_letter
            for cell in col:
                try:
                    val = str(cell.value) if cell.value else ""
                    max_len = max(max_len, len(val))
                except Exception:
                    pass
            ws.column_dimensions[col_letter].width = min(max_len + 4, 50)

    tmp = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
    wb.save(tmp.name)
    tmp.close()
    return tmp.name


# ── Review workflow ─────────────────────────────────────────────────────

VALID_REVIEW_STATUSES = {"pending", "reviewed", "approved"}


def review_statement(
    db: Session, statement_id: int, review_status: str,
    reviewer_id: str | None = None, review_notes: str | None = None,
) -> FinancialStatement:
    if review_status not in VALID_REVIEW_STATUSES:
        raise bad_request(f"Invalid review_status. Must be one of: {VALID_REVIEW_STATUSES}")
    stmt = get_statement(db, statement_id)
    if stmt.locked:
        raise bad_request("Statement is locked and cannot be modified")
    stmt.review_status = review_status
    stmt.reviewer_id = reviewer_id
    stmt.review_notes = review_notes
    db.commit()
    db.refresh(stmt)
    return stmt


def lock_statement(db: Session, statement_id: int) -> FinancialStatement:
    stmt = get_statement(db, statement_id)
    stmt.locked = True
    stmt.review_status = "approved"
    db.commit()
    db.refresh(stmt)
    return stmt


def edit_line_item(
    db: Session, line_item_id: int,
    edited_label: str | None = None, edited_value: float | None = None,
) -> LineItem:
    li = db.query(LineItem).filter(LineItem.id == line_item_id).first()
    if not li:
        raise not_found("Line item not found")

    stmt = db.query(FinancialStatement).filter(
        FinancialStatement.id == li.statement_id
    ).first()
    if stmt and stmt.locked:
        raise bad_request("Statement is locked and cannot be modified")

    # Log changes
    if edited_label is not None and edited_label != li.edited_label:
        db.add(EditLog(
            line_item_id=li.id, field="label",
            old_value=li.edited_label or li.label, new_value=edited_label,
        ))
        li.edited_label = edited_label

    if edited_value is not None and edited_value != li.edited_value:
        db.add(EditLog(
            line_item_id=li.id, field="value",
            old_value=str(li.edited_value if li.edited_value is not None else li.value),
            new_value=str(edited_value),
        ))
        li.edited_value = edited_value

    if edited_label is not None or edited_value is not None:
        li.is_user_modified = True

    db.commit()
    db.refresh(li)
    return li


def get_edit_history(db: Session, line_item_id: int) -> list[EditLog]:
    return db.query(EditLog).filter(
        EditLog.line_item_id == line_item_id
    ).order_by(EditLog.created_at.desc()).all()


# ── Investment mapping ──────────────────────────────────────────────────

def map_statement_to_investment(
    db: Session, statement_id: int, investment_id: int,
    reporting_date: str | None = None, fiscal_period_label: str | None = None,
) -> FinancialStatement:
    stmt = get_statement(db, statement_id)
    inv = db.query(Investment).filter(Investment.id == investment_id).first()
    if not inv:
        raise not_found("Investment not found")
    stmt.investment_id = investment_id
    stmt.reporting_date = reporting_date
    stmt.fiscal_period_label = fiscal_period_label
    db.commit()
    db.refresh(stmt)
    return stmt


def get_investment_financials(db: Session, investment_id: int) -> list[FinancialStatement]:
    return db.query(FinancialStatement).filter(
        FinancialStatement.investment_id == investment_id
    ).order_by(
        FinancialStatement.reporting_date.desc(),
        FinancialStatement.statement_type,
    ).all()


def suggest_investment_mapping(db: Session, statement_id: int) -> dict:
    """Auto-suggest investment mapping based on document metadata."""
    stmt = get_statement(db, statement_id)
    doc = db.query(Document).filter(Document.id == stmt.document_id).first()
    if not doc:
        return {"suggestions": []}

    suggestions = []
    # The document already belongs to an investment
    inv = db.query(Investment).filter(Investment.id == doc.investment_id).first()
    if inv:
        suggestions.append({
            "investment_id": inv.id,
            "investment_name": inv.investment_name,
            "confidence": "high",
            "reason": "Document belongs to this investment",
        })

    return {
        "suggestions": suggestions,
        "period": stmt.period,
        "period_end_date": stmt.period_end_date,
    }
